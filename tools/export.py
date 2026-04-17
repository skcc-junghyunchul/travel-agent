import json
import os
from datetime import datetime
from langchain_core.tools import tool
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
ACCENT_FONT = Font(color="FFFFFF", bold=True)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _header(ws, row, col, value, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill or HEADER_FILL
    cell.font = font or HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    return cell


def _cell(ws, row, col, value, bold=False, alt=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border
    if alt:
        cell.fill = ALT_FILL
    if bold:
        cell.font = Font(bold=True)
    return cell


def _autofit(ws, min_width=12, max_width=50):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_width), max_width)


@tool
def export_itinerary_to_excel(itinerary_json: str, output_path: str = "") -> str:
    """Export a travel itinerary to a formatted Excel file.

    The itinerary_json must be a JSON string with this structure:
    {
      "destination": "Paris, France",
      "departure_city": "New York, USA",
      "start_date": "2024-06-01",
      "end_date": "2024-06-08",
      "duration_days": 7,
      "travelers": 2,
      "age_groups": ["adult"],
      "preferences": ["culture", "food"],
      "outbound_flight": {"airline": "...", "flight_number": "...", "departure_airport": "...",
                          "arrival_airport": "...", "departure_time": "...", "arrival_time": "...",
                          "price_per_person": 850, "booking_reference": null},
      "return_flight": {...},
      "days": [
        {
          "day": 1, "date": "2024-06-02", "title": "Arrival Day",
          "accommodation": {"name": "...", "address": "...", "check_in": "...", "check_out": "...",
                            "price_per_night": 180, "booking_reference": null},
          "activities": [{"time": "10:00", "name": "...", "description": "...", "location": "...",
                          "duration": "2h", "cost": 25, "booking_reference": null}],
          "meals": [{"meal_type": "lunch", "restaurant": "...", "cuisine": "...", "address": "...",
                     "price_range": "$$", "booking_reference": null}],
          "transportation": [{"from_location": "...", "to_location": "...", "mode": "...",
                              "duration": "30min", "cost": 10}],
          "workout": {"name": "...", "workout_type": "running", "description": "...",
                      "location": "...", "distance_or_details": "3km loop"}
        }
      ],
      "cost_summary": {"flights": 1700, "accommodation": 1080, "activities": 300,
                       "meals": 420, "transportation": 150, "total": 3650}
    }

    Args:
        itinerary_json: Full itinerary as JSON string
        output_path: Where to save the file. Defaults to ./itinerary_<destination>_<date>.xlsx
    """
    try:
        data = json.loads(itinerary_json)
    except json.JSONDecodeError as e:
        return f"Invalid JSON: {str(e)}"

    destination_slug = data.get("destination", "trip").replace(", ", "_").replace(" ", "_")
    if not output_path:
        output_path = f"itinerary_{destination_slug}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()

    # --- Sheet 1: Overview ---
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov.row_dimensions[1].height = 40

    _header(ws_ov, 1, 1, "TRAVEL ITINERARY", fill=PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid"), font=Font(color="FFFFFF", bold=True, size=14))
    ws_ov.merge_cells("A1:D1")

    fields = [
        ("Destination", data.get("destination", "")),
        ("Departure City", data.get("departure_city", "")),
        ("Travel Dates", f"{data.get('start_date', '')} → {data.get('end_date', '')}"),
        ("Duration", f"{data.get('duration_days', '')} days"),
        ("Travelers", str(data.get("travelers", ""))),
        ("Age Groups", ", ".join(data.get("age_groups", []))),
        ("Preferences", ", ".join(data.get("preferences", []))),
    ]
    for i, (label, value) in enumerate(fields, 2):
        _cell(ws_ov, i, 1, label, bold=True, alt=(i % 2 == 0))
        _cell(ws_ov, i, 2, value, alt=(i % 2 == 0))

    cost = data.get("cost_summary", {})
    if cost:
        start_row = len(fields) + 3
        _header(ws_ov, start_row, 1, "Cost Summary")
        ws_ov.merge_cells(f"A{start_row}:B{start_row}")
        cost_items = [
            ("Flights", cost.get("flights", 0)),
            ("Accommodation", cost.get("accommodation", 0)),
            ("Activities", cost.get("activities", 0)),
            ("Meals", cost.get("meals", 0)),
            ("Transportation", cost.get("transportation", 0)),
            ("TOTAL", cost.get("total", 0)),
        ]
        for j, (label, amount) in enumerate(cost_items, start_row + 1):
            bold = label == "TOTAL"
            _cell(ws_ov, j, 1, label, bold=bold, alt=(j % 2 == 0))
            _cell(ws_ov, j, 2, f"${amount:,.2f}" if isinstance(amount, (int, float)) else amount, bold=bold, alt=(j % 2 == 0))

    _autofit(ws_ov)

    # --- Sheet 2: Day-by-Day Itinerary ---
    ws_days = wb.create_sheet("Daily Itinerary")
    headers = ["Day", "Date", "Time", "Type", "Name", "Description", "Location", "Duration/Details", "Cost (USD)", "Booking Ref"]
    for col, h in enumerate(headers, 1):
        _header(ws_days, 1, col, h)

    row = 2
    for day in data.get("days", []):
        day_label = f"Day {day.get('day', '')} — {day.get('title', '')}"
        _header(ws_days, row, 1, day_label, fill=ACCENT_FILL, font=ACCENT_FONT)
        ws_days.merge_cells(f"A{row}:J{row}")
        row += 1

        for act in day.get("activities", []):
            alt = row % 2 == 0
            for col, val in enumerate([
                day.get("day"), day.get("date"), act.get("time"), "Activity",
                act.get("name"), act.get("description"), act.get("location"),
                act.get("duration"), f"${act.get('cost', 0):.2f}", act.get("booking_reference", "—"),
            ], 1):
                _cell(ws_days, row, col, val, alt=alt)
            row += 1

        for meal in day.get("meals", []):
            alt = row % 2 == 0
            for col, val in enumerate([
                day.get("day"), day.get("date"), meal.get("meal_type", "").title(), "Meal",
                meal.get("restaurant"), meal.get("cuisine"), meal.get("address"),
                meal.get("price_range"), "", meal.get("booking_reference", "—"),
            ], 1):
                _cell(ws_days, row, col, val, alt=alt)
            row += 1

        for tr in day.get("transportation", []):
            alt = row % 2 == 0
            for col, val in enumerate([
                day.get("day"), day.get("date"), "", "Transport",
                tr.get("mode"), f"{tr.get('from_location')} → {tr.get('to_location')}",
                "", tr.get("duration"), f"${tr.get('cost', 0):.2f}", "—",
            ], 1):
                _cell(ws_days, row, col, val, alt=alt)
            row += 1

        wo = day.get("workout")
        if wo:
            alt = row % 2 == 0
            for col, val in enumerate([
                day.get("day"), day.get("date"), "", "Workout",
                wo.get("name"), wo.get("description"), wo.get("location"),
                wo.get("distance_or_details"), "Free", "—",
            ], 1):
                _cell(ws_days, row, col, val, alt=alt)
            row += 1

    _autofit(ws_days)

    # --- Sheet 3: Flights ---
    ws_fl = wb.create_sheet("Flights")
    fl_headers = ["Type", "Airline", "Flight No.", "From", "To", "Departure", "Arrival", "Price/Person", "Booking Ref"]
    for col, h in enumerate(fl_headers, 1):
        _header(ws_fl, 1, col, h)

    row = 2
    for ftype, fdata in [("Outbound", data.get("outbound_flight")), ("Return", data.get("return_flight"))]:
        if fdata:
            for col, val in enumerate([
                ftype, fdata.get("airline"), fdata.get("flight_number"),
                fdata.get("departure_airport"), fdata.get("arrival_airport"),
                fdata.get("departure_time"), fdata.get("arrival_time"),
                f"${fdata.get('price_per_person', 0):.2f}", fdata.get("booking_reference", "—"),
            ], 1):
                _cell(ws_fl, row, col, val, alt=row % 2 == 0)
            row += 1
    _autofit(ws_fl)

    # --- Sheet 4: Accommodation ---
    ws_ac = wb.create_sheet("Accommodation")
    ac_headers = ["Hotel / Property", "Address", "Check-in", "Check-out", "Nights", "Price/Night", "Total", "Booking Ref"]
    for col, h in enumerate(ac_headers, 1):
        _header(ws_ac, 1, col, h)

    seen = set()
    row = 2
    for day in data.get("days", []):
        acc = day.get("accommodation")
        if acc and acc.get("name") not in seen:
            seen.add(acc.get("name"))
            nights = (
                datetime.strptime(acc.get("check_out", ""), "%Y-%m-%d")
                - datetime.strptime(acc.get("check_in", ""), "%Y-%m-%d")
            ).days if acc.get("check_in") and acc.get("check_out") else ""
            total = nights * acc.get("price_per_night", 0) if isinstance(nights, int) else ""
            for col, val in enumerate([
                acc.get("name"), acc.get("address"), acc.get("check_in"), acc.get("check_out"),
                nights, f"${acc.get('price_per_night', 0):.2f}",
                f"${total:.2f}" if isinstance(total, float) else total,
                acc.get("booking_reference", "—"),
            ], 1):
                _cell(ws_ac, row, col, val, alt=row % 2 == 0)
            row += 1
    _autofit(ws_ac)

    # --- Sheet 5: Workout & Fitness ---
    ws_wo = wb.create_sheet("Workout Areas")
    wo_headers = ["Day", "Date", "Name", "Type", "Location", "Details"]
    for col, h in enumerate(wo_headers, 1):
        _header(ws_wo, 1, col, h)

    row = 2
    for day in data.get("days", []):
        wo = day.get("workout")
        if wo:
            for col, val in enumerate([
                day.get("day"), day.get("date"),
                wo.get("name"), wo.get("workout_type"), wo.get("location"), wo.get("distance_or_details"),
            ], 1):
                _cell(ws_wo, row, col, val, alt=row % 2 == 0)
            row += 1
    _autofit(ws_wo)

    wb.save(output_path)
    abs_path = os.path.abspath(output_path)
    return f"✅ Itinerary exported to Excel!\nFile: {abs_path}\nSheets: Overview, Daily Itinerary, Flights, Accommodation, Workout Areas"
